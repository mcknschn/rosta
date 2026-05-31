"""Brå (bra.se) — kriminalstatistik (delpoäng D, kategori trygghet).

Brå har inget publikt API men publicerar Sveriges officiella statistik (SOS) som fritt
vidareutnyttjningsbara Excel-tabeller (se bra.se/om-bra/om-webbplatsen/data-fran-bra.html).
Vi laddar ned och parsar de tabeller vars struktur är stabil och vars semantik entydigt
matchar en kanonisk indikator; råfilen cachas lokalt med manifest.

Implementerat:
  * Konstaterade fall av dödligt våld PER 100 000 invånare (Tabell 20)
      -> trygghet / grov_brottslighet / dodligt_vald (riktning down). Per-capita-talet
      (jämförbart över tid), inte råa antal.
  * NTU (Nationella trygghetsundersökningen), Tabellsamling 2007-2025:
      - sjalvrapporterad utsatthet för brott mot enskild person, andel av befolkningen
        (blad 3A, "Samtliga 16-84 år") -> trygghet / utsatthet_trygghet / brottsutsatthet
      - otrygghet vid utevistelse sent på kvällen i bostadsområdet, andel
        (blad 4A:1, "Samtliga 16-84 år") -> trygghet / utsatthet_trygghet / upplevd_otrygghet
      Båda riktning down (lägre = bättre). Endast den serie som är jämförbar under NUVARANDE
      NTU-metod tas med (brottsutsatthet: aggregatkategorin finns bara fr.o.m. 2016 enligt
      källfotnot; otrygghet: 2007-2016 är omräknade med ANNAN metod -> exkluderas, fr.o.m.
      2017 är nuvarande metod). Tidsbrytningen är källflaggad (asterisk i tabellen).

Återstår (separata Brå-produkter/tabeller, se docs/fas3_coverage.md): uppklaring/
handläggningstid, återfall i brott.
"""

from __future__ import annotations

import io
import json
from dataclasses import asdict
from typing import Any

import httpx
import openpyxl

from .base import RAW_DIR, Manifest, _safe

BASE = "https://bra.se"
# SOS får vidareutnyttjas fritt med källangivelse (bra.se/om-bra/om-webbplatsen/data-fran-bra).
LICENSE = "Sveriges officiella statistik (SOS) – fri vidareutnyttjning med källangivelse (källa: Brå)"
_UA = {"User-Agent": "rosta-datapipeline/0.1 (civic-tech; official swedish open data)"}

# Stabil nedladdnings-URL verifierad 2026-05-30. Brå:s download-id kan ändras vid ny publicering
# -> en HTTP-fel höjs då tydligt (raise_for_status) i stället för att tyst ge gammal data.
DODLIGT_VALD_URL = (
    f"{BASE}/download/18.5b3bbb9a19d24bdce4b157ff/1774872877489/Tabell%2020_2002-2025.xlsx"
)

# NTU-tabellsamlingen (xlsx). Download-id verifierat 2026-05-30; ändras vid ny publicering
# -> HTTP-fel höjs då (raise_for_status) i stället för att tyst ge gammal data.
NTU_URL = (
    f"{BASE}/download/18.73837f1119e639446eb674d/1779977985828/"
    "Tabellsamling%20NTU%202007-2025.xlsx"
)

# NTU-serier vi tar in. Varje serie pekar ut EN rad i ETT blad via dess radetikett (kol A[/B]).
# min_year skär bort de år som inte är jämförbara under nuvarande NTU-metod (källflaggat med *):
#   3A  brottsutsatthet: aggregatet "Brott mot enskild person" finns bara fr.o.m. 2016 (fotnot 2).
#   4A:1 upplevd_otrygghet: 2007-2016 är omräknade med annan metod; nuvarande metod fr.o.m. 2017.
_NTU_SERIES = (
    {
        "sheet": "3A", "label_a": "Brott mot enskild person", "label_b": "Samtliga",
        "indicator": "brottsutsatthet", "min_year": 2016,
    },
    {
        "sheet": "4A.1", "label_a": "Samtliga 16-84 år", "label_b": None,
        "indicator": "upplevd_otrygghet", "min_year": 2017,
    },
)

# Kanoniska indikatorer denna modul levererar (för täcknings-gaten i tests/test_fas3_gate.py).
INDICATORS = ("dodligt_vald", "brottsutsatthet", "upplevd_otrygghet")


def _norm(s: Any) -> str:
    """Normaliserar en etikett för robust matchning: gemener, dash-varianter -> '-', en blank."""
    if s is None:
        return ""
    t = str(s).strip().lower()
    for dash in ("–", "—", "−"):  # en dash, em dash, minustecken
        t = t.replace(dash, "-")
    return " ".join(t.split())


def _as_year(cell: Any) -> int | None:
    """Tolkar en headercell som årtal. '2016' och '2016*' -> 2016; CI-rubriker/övrigt -> None."""
    if cell is None:
        return None
    t = str(cell).strip().rstrip("*").strip()
    if len(t) == 4 and t.isdigit():
        y = int(t)
        if 1990 <= y <= 2100:
            return y
    return None


def _ntu_header(ws: Any) -> tuple[int, dict[int, int]]:
    """Hittar årsheadern (raden med flest årtalsceller) -> (radindex, {kolumn -> år}).

    CI-kolumner ("Konfidensintervall ... 2024") räknas inte som år (hela strängen ≠ 4 siffror),
    så de utesluts automatiskt. Kräver minst 3 årskolumner — annars hård fail (struktur ändrad).
    """
    best_row: int | None = None
    best_map: dict[int, int] = {}
    for r in range(1, min(ws.max_row, 12) + 1):
        m: dict[int, int] = {}
        for c in range(1, ws.max_column + 1):
            y = _as_year(ws.cell(r, c).value)
            if y is not None:
                m[c] = y
        if len(m) > len(best_map):
            best_map, best_row = m, r
    if best_row is None or len(best_map) < 3:
        raise ValueError(f"NTU: hittar ingen års-header i blad {ws.title!r}")
    return best_row, best_map


def _ntu_find_row(ws: Any, header_row: int, label_a: str, label_b: str | None) -> int:
    """Hittar EXAKT en datarad vars kol A == label_a (och ev. kol B börjar med label_b)."""
    na, nb = _norm(label_a), (_norm(label_b) if label_b else None)
    hits = [
        r for r in range(header_row + 1, ws.max_row + 1)
        if _norm(ws.cell(r, 1).value) == na
        and (nb is None or _norm(ws.cell(r, 2).value).startswith(nb))
    ]
    if len(hits) != 1:
        raise ValueError(
            f"NTU blad {ws.title!r}: förväntade exakt 1 rad för {label_a!r}/{label_b!r}, fick {hits}"
        )
    return hits[0]


def _ntu_headline_series(ws: Any, label_a: str, label_b: str | None, min_year: int) -> dict[int, float]:
    """Plockar {år -> värde} för rubrikraden, fr.o.m. min_year. Icke-numeriskt ('..','.') hoppas.

    Kräver minst 2 numeriska årsvärden — annars hård fail (hellre tomt än tyst fel serie)."""
    hrow, col_year = _ntu_header(ws)
    drow = _ntu_find_row(ws, hrow, label_a, label_b)
    out: dict[int, float] = {}
    for col, year in col_year.items():
        if year < min_year:
            continue
        v = ws.cell(drow, col).value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            out[year] = round(float(v), 2)
    if len(out) < 2:
        raise ValueError(
            f"NTU blad {ws.title!r}: <2 numeriska årsvärden för {label_a!r} (min_year={min_year})"
        )
    return out


def _ntu_rows_from_workbook(wb: Any) -> list[dict[str, Any]]:
    """Bygger observations-rader ur en (redan inläst) NTU-arbetsbok. Nätverksfri -> testbar."""
    rows: list[dict[str, Any]] = []
    for s in _NTU_SERIES:
        ws = wb[s["sheet"]]
        series = _ntu_headline_series(ws, s["label_a"], s["label_b"], s["min_year"])
        sheet_tag = s["sheet"].lower().replace(".", "")  # '4A.1' -> '4a1'
        for year, val in sorted(series.items()):
            rows.append({
                "id": f"obs:bra:{s['indicator']}:{year}",
                "category": "trygghet", "submeasure": "utsatthet_trygghet",
                "indicator": s["indicator"], "period": str(year),
                "value": val, "unit": "% (andel av befolkningen 16–84 år)",
                "geography": "Riket", "source_ref": f"bra:ntu_{sheet_tag}:{year}",
            })
    return rows


def fetch_ntu(retrieved_at: str) -> list[dict[str, Any]]:
    """NTU Tabellsamling: utsatthet (3A) + otrygghet (4A:1), Samtliga 16-84 år -> observations."""
    with _client() as c:
        resp = c.get(NTU_URL)
        resp.raise_for_status()
        content = resp.content
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    rows = _ntu_rows_from_workbook(wb)
    _cache("ntu_tabellsamling", retrieved_at, NTU_URL, content, {"rows": len(rows)}, len(rows))
    return rows


def _client() -> httpx.Client:
    return httpx.Client(timeout=60, headers=_UA, follow_redirects=True)


def _cache(dataset_id: str, retrieved_at: str, url: str, content: bytes,
           payload: Any, rows: int) -> None:
    base = RAW_DIR / "bra" / _safe(dataset_id)
    base.mkdir(parents=True, exist_ok=True)
    (base / f"{_safe(retrieved_at)}.xlsx").write_bytes(content)  # rå Excel (reproducerbarhet)
    man = Manifest(
        source="bra", dataset_id=dataset_id, url=url, query="xlsx download",
        retrieved_at=retrieved_at, license=LICENSE, row_count=rows,
    )
    path = base / f"{_safe(retrieved_at)}.json"
    tmp = path.with_name(path.name + ".tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump({"manifest": asdict(man), "payload": payload}, fh, ensure_ascii=False)
    tmp.replace(path)


def _dodligt_rows_from_workbook(wb: Any) -> list[dict[str, Any]]:
    """Parsar Tabell 20-arbetsboken -> observations-rader. Nätverksfri -> testbar."""
    ws = wb["Statistik"]  # rad 1 = titel, rad 2 = rubriker, rad 3+ = (år, antal, per 100 000)
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        cells = list(values) + [None, None, None]
        year, _count, per_100k = cells[0], cells[1], cells[2]
        # Hoppa rader utan numeriskt år eller numeriskt värde. SOS markerar saknad data med
        # strängen '.' ("Punkt (.) innebär att adekvat data saknas"); float('.') skulle krascha
        # hela ingesten, så icke-numeriska värden behandlas som saknade.
        if not isinstance(year, (int, float)) or not isinstance(per_100k, (int, float)):
            continue
        # bool är en subklass till int -> uteslut explicit (en TRUE/FALSE-cell är inte ett år).
        if isinstance(year, bool) or isinstance(per_100k, bool):
            continue
        y = int(year)
        rows.append({
            "id": f"obs:bra:dodligt_vald:{y}",
            "category": "trygghet", "submeasure": "grov_brottslighet",
            "indicator": "dodligt_vald", "period": str(y),
            "value": float(per_100k), "unit": "per 100 000 inv.",
            "geography": "Riket", "source_ref": f"bra:dodligt_vald_tabell20:{y}",
        })
    return rows


def fetch_dodligt_vald(retrieved_at: str) -> list[dict[str, Any]]:
    """Tabell 20: konstaterade fall av dödligt våld per 100 000 inv. -> observations (Riket)."""
    with _client() as c:
        resp = c.get(DODLIGT_VALD_URL)
        resp.raise_for_status()
        content = resp.content
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    rows = _dodligt_rows_from_workbook(wb)
    _cache("dodligt_vald", retrieved_at, DODLIGT_VALD_URL, content, {"rows": len(rows)}, len(rows))
    return rows
