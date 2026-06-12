"""Reproducerar och AUDITERAR config/asyl_handlaggningstid.yaml mot Migrationsverkets xlsx.

Migrationsverket publicerar "Avgjorda asylärenden" som en xlsx per år. Den genomsnittliga
handläggningstiden för reguljära förstagångsärenden ligger i bladet "Totalt, förstagångsärenden"
(2021: "Totalt, första ansökan"), i deltabellen "Asyl" — den FÖRSTA "Totalt"-raden — i kolumnen
"Handläggningstid, dagar". De efterföljande "Totalt"-raderna (massflyktsdirektivet / ukrainska
medborgare) EXKLUDERAS medvetet (near-automatisk EU-process, ej svensk handläggningseffektivitet).

Det här verktyget hämtar varje års pinnade käll-xlsx (config: years[].source), extraherar den första
Totalt-radens handläggningstid och korsverifierar mot configens värde.

Kör manuellt (ej en pipeline-/test-dependens — kräver nätverk + openpyxl):

    python -m pipeline.tools.asyl_handlaggningstid_verify

Exit 0 = configen matchar de officiella filerna; exit 1 = avvikelse (värde ändrat i källan, ny
datapunkt, eller död URL). Detta är transkriberingens revisionsspår — ingen runtime-xlsx-parser körs
i själva pipelinen (migrationsverket.py läser bara den pinnade configen).
"""

from __future__ import annotations

import io

import httpx
import openpyxl

from .. import config

_UA = {"User-Agent": "rosta-datapipeline/0.1 (civic-tech; official swedish open data)"}
_TOL = 0  # handläggningstiden publiceras som heltal dagar -> exakt match krävs


def _first_application_sheet(wb: openpyxl.Workbook) -> str:
    """Bladet med förstagångsärenden om asyl (ej EKB, förlängning eller medborgarskap)."""
    for name in wb.sheetnames:
        n = name.lower()
        first = "första" in n or "forsta" in n or "förstag" in n or "forstag" in n
        excl = any(x in n for x in ("ekb", "medborgar", "förläng", "forläng", "forlang"))
        if first and not excl:
            return name
    raise ValueError(f"Hittar inget förstagångs-/första-ansökan-blad i {wb.sheetnames}")


def extract_handlaggningstid(xlsx_bytes: bytes) -> int:
    """Plockar deltabellen Asyls (FÖRSTA Totalt-radens) handläggningstid (dagar) ur en års-xlsx."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[_first_application_sheet(wb)]
    rows = list(ws.iter_rows(values_only=True))
    hcol = None
    header_i = None
    for i, row in enumerate(rows):
        cells = [str(c) if c is not None else "" for c in row]
        joined = " ".join(cells).lower()
        if "handl" in joined and "dagar" in joined:
            header_i = i
            for j, c in enumerate(cells):
                if "handl" in c.lower():
                    hcol = j
            break
    if header_i is None or hcol is None:
        raise ValueError("Hittar ingen 'Handläggningstid, dagar'-kolumn (bladlayout ändrad?)")
    for row in rows[header_i + 1:]:
        c0 = str(row[0]).strip() if row[0] is not None else ""
        if c0.lower() == "totalt":  # FÖRSTA Totalt = deltabellen Asyl
            val = row[hcol] if hcol < len(row) else None
            if val is None:
                raise ValueError("Tom handläggningstid på Totalt-raden")
            return int(round(float(val)))
    raise ValueError("Hittar ingen 'Totalt'-rad efter rubriken (bladlayout ändrad?)")


def main() -> int:
    cfg = config.asyl_handlaggningstid()
    ok = True
    with httpx.Client(timeout=60, headers=_UA, follow_redirects=True) as c:
        for year, entry in sorted(cfg["years"].items()):
            url = entry["source"]
            try:
                got = extract_handlaggningstid(c.get(url).content)
            except Exception as e:  # noqa: BLE001 — verktyget ska rapportera, ej krascha
                print(f"  {year}: FEL vid hämtning/läsning: {e!r}")
                ok = False
                continue
            want = int(entry["value"])
            mark = "OK" if abs(got - want) <= _TOL else "AVVIKER"
            if mark != "OK":
                ok = False
            print(f"  {year}: config={want:>4}  källa={got:>4}  -> {mark}")
    print("KLART:", "configen matchar de officiella filerna" if ok else "AVVIKELSE — se ovan")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
